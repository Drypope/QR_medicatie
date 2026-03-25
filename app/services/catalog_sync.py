from __future__ import annotations

import hashlib
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import MedicationCatalog

REQUIRED_COLUMNS = {"product_class", "product_name", "unique_id"}


def _file_sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _normalize_unique_id(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        return text[:-2]
    return text


def sync_catalog(session: Session, catalog_path: Path) -> int:
    if not catalog_path.exists():
        raise FileNotFoundError(f"Catalog file not found: {catalog_path}")

    workbook = load_workbook(catalog_path, read_only=True, data_only=True)
    if "catalog" not in workbook.sheetnames:
        raise ValueError("Missing required 'catalog' sheet")

    sheet = workbook["catalog"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return 0

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    header_to_idx = {h: i for i, h in enumerate(headers) if h}
    missing = REQUIRED_COLUMNS - set(header_to_idx)
    if missing:
        raise ValueError(f"Missing required catalog columns: {sorted(missing)}")

    source_hash = _file_sha256(catalog_path)
    synced_at = datetime.now(timezone.utc)

    existing_by_name = {
        row.product_name: row
        for row in session.scalars(select(MedicationCatalog)).all()
    }

    upsert_count = 0
    seen_product_names: set[str] = set()

    for raw in rows[1:]:
        if raw is None:
            continue
        product_name = str(raw[header_to_idx["product_name"]] or "").strip()
        product_class = str(raw[header_to_idx["product_class"]] or "").strip()
        unique_id = _normalize_unique_id(raw[header_to_idx["unique_id"]])
        if not (product_name and product_class and unique_id):
            continue

        seen_product_names.add(product_name)

        sort_order = int(raw[header_to_idx.get("sort_order", -1)] or 0) if "sort_order" in header_to_idx else 0
        is_active_cell = raw[header_to_idx.get("is_active", -1)] if "is_active" in header_to_idx else True
        is_active = bool(is_active_cell) if is_active_cell is not None else True

        existing = existing_by_name.get(product_name)
        if existing is None:
            existing = MedicationCatalog(product_name=product_name)
            session.add(existing)

        existing.product_name = product_name
        existing.product_class = product_class
        existing.unique_id = unique_id
        existing.sort_order = sort_order
        existing.is_active = is_active
        existing.source_hash = source_hash
        existing.last_synced_at = synced_at
        upsert_count += 1

    for existing in existing_by_name.values():
        if existing.product_name not in seen_product_names:
            existing.is_active = False
            existing.last_synced_at = synced_at

    session.commit()
    return upsert_count
